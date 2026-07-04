from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.models.journey import Journey
from app.models.participant import Participant
from app.models.prediction import Prediction, PredictionPick
from app.models.race import Race
from app.services.tips_excel_generator import TipsGenerationError, generate_tips_excel

TEMPLATE = Path("templates_excel") / "Tips _Ejemplo.xlsx"


def make_journey(venue: str, n_races: int, distinct_pick1: int = 4) -> Journey:
    journey = Journey(date=date(2026, 7, 2), venue=venue, theme="black", season_year=2026)
    for race_number in range(1, n_races + 1):
        race = Race(race_number=race_number)
        race.participants = [Participant(number=i, horse_name=f"HORSE R{race_number} N{i}") for i in range(1, 13)]
        journey.races.append(race)
    for index in range(8):
        prediction = Prediction()
        prediction.picks = [
            PredictionPick(
                race_number=race_number,
                pick_1=(index % distinct_pick1) + 1,
                pick_2=(index % distinct_pick1) + 5,
                pick_3=(index % distinct_pick1) + 7,
            )
            for race_number in range(1, n_races + 1)
        ]
        journey.predictions.append(prediction)
    return journey


@pytest.mark.parametrize(
    ("venue", "sheet", "n_races"),
    [
        ("HZ_MADRID", "HZ", 8),
        ("SAN_SEBASTIAN", "SS", 7),
        ("DOS_HERMANAS", "DH", 5),
        ("SANLUCAR", "SL", 7),
        ("PINEDA", "PIN", 5),
        ("HZ_NOCTURNAS", "NOCTURNAS", 5),
    ],
)
def test_generates_venue_sheet_with_template_format(tmp_path, venue, sheet, n_races) -> None:
    output = tmp_path / "tips.xlsx"
    generate_tips_excel(make_journey(venue, n_races), output, TEMPLATE)

    wb = load_workbook(output)
    assert wb.sheetnames == [sheet]
    ws = wb.active
    merges = {str(m) for m in ws.merged_cells.ranges}
    for group in range(1, n_races + 1):
        name_col = 2 + 3 * (group - 1)
        assert ws.cell(row=2, column=name_col).value == f"{group}ª CARRERA"
    # Datos: caballos con nombre oficial y votos agregados de pick_1.
    assert ws.cell(row=3, column=2).value == "HORSE R1 N1 (1)"
    assert ws.cell(row=3, column=3).value == 2  # 8 especialistas / 4 pick_1 distintos
    # Fills alternos conservados en la zona de datos.
    assert ws.cell(row=4, column=2).fill.fill_type == "solid"
    # Footer con etiquetas sin valores de ejemplo.
    assert ws.cell(row=10, column=2).value == "FAVORITO DE LA JORNADA:"
    assert ws.cell(row=11, column=2).value == "ENTRENADOR DE LA JORNADA:"
    assert ws.cell(row=12, column=2).value == "JOCKEY DE LA JORNADA:"
    assert {"B10:E10", "B11:E11", "B12:E12"} <= merges


def test_blanks_surplus_race_groups(tmp_path) -> None:
    output = tmp_path / "tips.xlsx"
    generate_tips_excel(make_journey("HZ_MADRID", 7), output, TEMPLATE)

    ws = load_workbook(output).active
    merges = {str(m) for m in ws.merged_cells.ranges}
    assert ws.cell(row=2, column=23).value is None  # W2 (grupo 8) limpiado
    assert "W2:X2" not in merges


def test_extends_race_groups_beyond_template(tmp_path) -> None:
    output = tmp_path / "tips.xlsx"
    generate_tips_excel(make_journey("HZ_NOCTURNAS", 6), output, TEMPLATE)

    ws = load_workbook(output).active
    merges = {str(m) for m in ws.merged_cells.ranges}
    assert ws.cell(row=2, column=17).value == "6ª CARRERA"  # Q2
    assert "Q2:R2" in merges
    assert ws.cell(row=3, column=17).value == "HORSE R6 N1 (1)"


def test_overflow_shifts_footer_and_keeps_labels(tmp_path) -> None:
    output = tmp_path / "tips.xlsx"
    generate_tips_excel(make_journey("DOS_HERMANAS", 5, distinct_pick1=8), output, TEMPLATE)

    ws = load_workbook(output).active
    merges = {str(m) for m in ws.merged_cells.ranges}
    assert ws.cell(row=10, column=2).value == "HORSE R1 N8 (8)"  # 8 filas de datos
    assert ws.cell(row=12, column=2).value == "FAVORITO DE LA JORNADA:"
    assert {"B12:E12", "B13:E13", "B14:E14"} <= merges
    assert "B10:E10" not in merges


def test_missing_template_raises(tmp_path) -> None:
    with pytest.raises(TipsGenerationError):
        generate_tips_excel(make_journey("HZ_MADRID", 5), tmp_path / "out.xlsx", tmp_path / "missing.xlsx")


def test_unknown_venue_raises(tmp_path) -> None:
    with pytest.raises(TipsGenerationError):
        generate_tips_excel(make_journey("SEDE_INEXISTENTE", 5), tmp_path / "out.xlsx", TEMPLATE)


def test_pick_outside_partant_blocks_generation(tmp_path) -> None:
    journey = make_journey("HZ_MADRID", 5)
    journey.predictions[0].picks[0].pick_1 = 99
    with pytest.raises(TipsGenerationError):
        generate_tips_excel(journey, tmp_path / "out.xlsx", TEMPLATE)
