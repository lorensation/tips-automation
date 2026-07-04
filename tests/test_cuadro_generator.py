import shutil
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.config import Settings
from app.enums import SPECIALIST_NAMES
from app.models.journey import Journey
from app.models.participant import Participant
from app.models.prediction import Prediction, PredictionPick
from app.models.race import Race
from app.models.specialist import Specialist
from app.services import cuadro_excel_generator
from app.services.cuadro_excel_generator import CuadroGenerationError, append_journey_sheet

REAL_PRONOS = Path("templates_excel") / "PRONOS 2026.xlsx"


def make_settings(tmp_path: Path) -> Settings:
    return Settings(
        pronos_file_path=tmp_path / "PRONOS 2026.xlsx",
        pronos_backup_dir=tmp_path / "backups",
    )


def make_journey(venue: str, theme: str, journey_date: date, n_races: int, picks: tuple[int, int, int] | None = None) -> Journey:
    journey = Journey(date=journey_date, venue=venue, theme=theme, season_year=journey_date.year)
    for race_number in range(1, n_races + 1):
        race = Race(race_number=race_number)
        race.participants = [
            Participant(number=i, horse_name=f"CAB R{race_number} N{i}", is_active=True) for i in range(1, 13)
        ]
        journey.races.append(race)
    for index, name in enumerate(SPECIALIST_NAMES):
        prediction = Prediction()
        prediction.specialist = Specialist(name=name, display_order=index + 1)
        values = picks or ((index % 4) + 1, (index % 4) + 5, (index % 4) + 7)
        prediction.picks = [
            PredictionPick(race_number=race_number, pick_1=values[0], pick_2=values[1], pick_3=values[2])
            for race_number in range(1, n_races + 1)
        ]
        journey.predictions.append(prediction)
    return journey


def test_bootstrap_creates_accumulative_with_hidden_template(tmp_path) -> None:
    settings = make_settings(tmp_path)
    journey = make_journey("HZ_NOCTURNAS", "black", date(2026, 7, 2), 5)

    path, sheet_name, replaced = append_journey_sheet(journey, settings)

    assert path.exists() and not replaced and sheet_name == "PRO_02_07"
    wb = load_workbook(path)
    assert wb["_PLANTILLA"].sheet_state == "hidden"
    ws = wb[sheet_name]
    assert ws["B17"].value == "ESPECIALISTAS"
    assert ws["A16"].value == "HZ_NOCTURNAS · 2026-07-02"
    assert ws["D17"].fill.fgColor.rgb == "FF262626"  # paleta negra nocturnas
    assert ws["C19"].value == "EMILIO VILLAVERDE"
    assert ws["D19"].value == "CAB R1 N1 (1)"
    assert [ws.cell(row=r, column=5).value for r in (19, 20, 21)] == [1, 5, 7]
    assert ws["D43"].value  # consenso presente
    assert len(ws._images) == 1  # logo re-insertado


def test_exact_cell_mapping_for_first_specialist_and_consensus(tmp_path) -> None:
    settings = make_settings(tmp_path)
    journey = make_journey("SAN_SEBASTIAN", "green", date(2026, 7, 5), 5, picks=(5, 3, 2))

    path, sheet_name, _ = append_journey_sheet(journey, settings)

    ws = load_workbook(path)[sheet_name]
    # EMILIO VILLAVERDE, carrera 1, picks [5, 3, 2]:
    assert ws["B19"].value == "1º"
    assert ws["C19"].value == "EMILIO VILLAVERDE"
    assert ws["D19"].value == "CAB R1 N5 (5)"  # nombre oficial del partant, no del texto
    assert ws["E19"].value == 5
    assert ws["E20"].value == 3
    assert ws["E21"].value == 2
    # El formato "5-3-2" SOLO aparece en la fila de consenso.
    assert ws["D43"].value == "5-3-2"
    for row in range(19, 43):
        for col in range(4, 14):
            assert ws.cell(row=row, column=col).value != "5-3-2"
    # Cabeceras correctas, sin C1/C2.
    headers = [ws.cell(row=17, column=4 + 2 * i).value for i in range(5)]
    assert headers == ["1ª CARRERA", "2ª CARRERA", "3ª CARRERA", "4ª CARRERA", "5ª CARRERA"]
    # Título fusionado hasta la última columna de carrera (M para 5 carreras).
    assert "B13:M15" in {str(m) for m in ws.merged_cells.ranges}


def test_append_on_real_file_keeps_history_and_positions_sheet(tmp_path) -> None:
    settings = make_settings(tmp_path)
    shutil.copy2(REAL_PRONOS, settings.pronos_file_path)
    before = load_workbook(settings.pronos_file_path)
    pre_names = list(before.sheetnames)
    pre_d19 = before["PRO_02_07"]["D19"].value
    before.close()

    journey = make_journey("SAN_SEBASTIAN", "green", date(2026, 7, 10), 7)
    _, sheet_name, replaced = append_journey_sheet(journey, settings)

    assert sheet_name == "PRO_10_07" and not replaced
    wb = load_workbook(settings.pronos_file_path)
    for name in pre_names:
        assert name in wb.sheetnames
    assert wb["PRO_02_07"]["D19"].value == pre_d19  # hoja histórica intacta
    assert wb.sheetnames.index(sheet_name) < wb.sheetnames.index("ANUAL 2026")
    ws = wb[sheet_name]
    assert ws.cell(row=17, column=16).value == "7ª CARRERA"  # par extra (col P)
    assert ws["D17"].fill.fgColor.rgb == "FF548235"  # paleta verde
    assert (settings.pronos_backup_dir).exists()
    assert list(settings.pronos_backup_dir.glob("*.xlsx"))


def test_regeneration_replaces_sheet_without_duplicates(tmp_path) -> None:
    settings = make_settings(tmp_path)
    shutil.copy2(REAL_PRONOS, settings.pronos_file_path)
    journey = make_journey("SAN_SEBASTIAN", "green", date(2026, 7, 10), 7)

    append_journey_sheet(journey, settings)
    _, sheet_name, replaced = append_journey_sheet(journey, settings)

    assert replaced
    wb = load_workbook(settings.pronos_file_path)
    assert wb.sheetnames.count(sheet_name) == 1


def test_failed_verification_leaves_original_untouched(tmp_path, monkeypatch) -> None:
    settings = make_settings(tmp_path)
    shutil.copy2(REAL_PRONOS, settings.pronos_file_path)
    before = load_workbook(settings.pronos_file_path)
    pre_names = list(before.sheetnames)
    before.close()

    def broken_verify(path, expected_sheets, new_sheet):
        raise CuadroGenerationError("verificación forzada a fallar")

    monkeypatch.setattr(cuadro_excel_generator, "_verify", broken_verify)
    journey = make_journey("HZ_MADRID", "blue", date(2026, 7, 24), 6)

    with pytest.raises(CuadroGenerationError):
        append_journey_sheet(journey, settings)

    wb = load_workbook(settings.pronos_file_path)
    assert list(wb.sheetnames) == pre_names  # el original no se tocó
    assert not list(tmp_path.glob("*.tmp.xlsx"))  # temporal limpiado


def test_incomplete_predictions_block_generation(tmp_path) -> None:
    settings = make_settings(tmp_path)
    journey = make_journey("HZ_MADRID", "blue", date(2026, 7, 24), 5)
    journey.predictions.pop()

    with pytest.raises(CuadroGenerationError):
        append_journey_sheet(journey, settings)


def test_invalid_pick_leaves_accumulative_untouched(tmp_path) -> None:
    settings = make_settings(tmp_path)
    shutil.copy2(REAL_PRONOS, settings.pronos_file_path)
    original_bytes = settings.pronos_file_path.read_bytes()
    journey = make_journey("HZ_MADRID", "blue", date(2026, 7, 24), 5)
    journey.predictions[0].picks[0].pick_1 = 99  # no existe en el partant

    with pytest.raises(CuadroGenerationError):
        append_journey_sheet(journey, settings)

    assert settings.pronos_file_path.read_bytes() == original_bytes
    assert not settings.pronos_backup_dir.exists() or not list(settings.pronos_backup_dir.iterdir())


def test_duplicate_picks_leave_accumulative_untouched(tmp_path) -> None:
    settings = make_settings(tmp_path)
    shutil.copy2(REAL_PRONOS, settings.pronos_file_path)
    original_bytes = settings.pronos_file_path.read_bytes()
    journey = make_journey("SANLUCAR", "yellow", date(2026, 7, 24), 5)
    journey.predictions[2].picks[1].pick_2 = journey.predictions[2].picks[1].pick_1  # duplicado

    with pytest.raises(CuadroGenerationError):
        append_journey_sheet(journey, settings)

    assert settings.pronos_file_path.read_bytes() == original_bytes
