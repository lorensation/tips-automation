"""Añade la hoja de una jornada al Excel acumulativo del cuadro de pronósticos.

El acumulativo es el archivo corporativo "PRONOS 2026.xlsx": nunca se recrea y
sus hojas históricas no se tocan. Cada escritura hace backup previo, guarda en
un temporal, verifica el resultado y solo entonces reemplaza el archivo.

La hoja nueva se copia de una hoja plantilla oculta "_PLANTILla" dentro del
propio acumulativo (creada una única vez desde la plantilla de ejemplo);
`copy_worksheet` preserva estilos, merges y dimensiones pero no imágenes, por
lo que el logo se re-inserta desde la plantilla.
"""

import logging
import os
import shutil
import zipfile
from copy import copy, deepcopy
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.config import Settings
from app.models.journey import Journey
from app.services.consensus import PickRow, calculate_consensus
from app.utils import utcnow

logger = logging.getLogger(__name__)

PLANTILLA_SHEET = "_PLANTILLA"
ANNUAL_SHEET = "ANUAL 2026"
BACKUP_RETENTION = 20

TITLE_TOP, TITLE_BOTTOM = 13, 15
SUBTITLE_ROW = 16
HEADER_TOP, HEADER_BOTTOM = 17, 18
SPEC_FIRST_ROW = 19
SPEC_BLOCK_ROWS = 3
N_SPECIALIST_BLOCKS = 8
CONSENSUS_ROW = 43
FIRST_PAIR_COL = 4  # columna D
TEMPLATE_IMAGE_ENTRY = "xl/media/image1.jpeg"
_VIRGIN_CELL = (1, 40)

CUADRO_PALETTES = {
    "blue": {"dark": "FF2E74B5", "light": "FFDCE6F1"},
    "black": {"dark": "FF262626", "light": "FFD9D9D9"},
    "green": {"dark": "FF548235", "light": "FFE2EFDA"},
    "orange": {"dark": "FFC55A11", "light": "FFFBE5D6"},
    "yellow": {"dark": "FFBF8F00", "light": "FFFFF2CC"},
    "purple": {"dark": "FF7030A0", "light": "FFE6DFEC"},
}


class CuadroGenerationError(Exception):
    pass


def append_journey_sheet(journey: Journey, settings: Settings) -> tuple[Path, str, bool]:
    """Devuelve (ruta del acumulativo, nombre de hoja, si se reemplazó una hoja previa)."""
    acc_path = settings.pronos_file_path
    template_path = settings.templates_excel_dir / settings.cuadro_template_filename
    if not template_path.exists():
        raise CuadroGenerationError(f"No se encuentra la plantilla del cuadro: {template_path}")
    n_races = len(journey.races)
    if n_races == 0:
        raise CuadroGenerationError("La jornada no tiene carreras.")

    bootstrapped = _bootstrap_if_missing(acc_path, template_path)
    backup_path = _backup(acc_path, settings.pronos_backup_dir)

    wb = load_workbook(acc_path)
    pre_sheets = set(wb.sheetnames)
    plantilla = _ensure_plantilla_sheet(wb, template_path)

    sheet_name = _sheet_name(journey, settings)
    replaced = sheet_name in wb.sheetnames
    if replaced:
        wb.remove(wb[sheet_name])

    new_ws = wb.copy_worksheet(plantilla)
    new_ws.title = sheet_name
    new_ws.sheet_state = "visible"
    _position_before_annual(wb, sheet_name)

    _adapt_race_pairs(new_ws, n_races)
    _fill_sheet(new_ws, journey, settings)
    _recolor(new_ws, journey.theme, n_races)
    _attach_logo(new_ws, template_path)

    expected_sheets = (pre_sheets - {sheet_name}) | {sheet_name, PLANTILLA_SHEET}
    tmp_path = acc_path.with_name(acc_path.stem + ".tmp.xlsx")
    try:
        wb.save(tmp_path)
        _verify(tmp_path, expected_sheets, sheet_name)
        os.replace(tmp_path, acc_path)
    except Exception as exc:
        tmp_path.unlink(missing_ok=True)
        if not _is_loadable(acc_path):
            shutil.copy2(backup_path, acc_path)
            logger.error("Acumulativo restaurado desde backup %s", backup_path)
        raise CuadroGenerationError(f"Fallo al escribir el acumulativo: {exc}") from exc

    logger.info(
        "Hoja %s %s en %s (backup: %s, bootstrap: %s)",
        sheet_name, "reemplazada" if replaced else "añadida", acc_path, backup_path, bootstrapped,
    )
    return acc_path, sheet_name, replaced


def _sheet_name(journey: Journey, settings: Settings) -> str:
    return settings.cuadro_sheet_name_pattern.format(
        dd=f"{journey.date.day:02d}",
        mm=f"{journey.date.month:02d}",
        date=journey.date.isoformat(),
        venue=journey.venue,
    )


def _bootstrap_if_missing(acc_path: Path, template_path: Path) -> bool:
    if acc_path.exists():
        return False
    acc_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, acc_path)
    wb = load_workbook(acc_path)
    ws = wb.active
    ws.title = PLANTILLA_SHEET
    ws.sheet_state = "hidden"
    # Un workbook necesita al menos una hoja visible.
    annual = wb.create_sheet(ANNUAL_SHEET)
    annual.sheet_state = "visible"
    wb.save(acc_path)
    logger.warning(
        "Acumulativo no encontrado: creado desde plantilla en %s. "
        "Coloca ahí el archivo real descargado de SharePoint para acumular sobre él.",
        acc_path,
    )
    return True


def _backup(acc_path: Path, backup_dir: Path) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = utcnow().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{acc_path.stem}.backup_{timestamp}{acc_path.suffix}"
    shutil.copy2(acc_path, backup_path)
    backups = sorted(backup_dir.glob(f"{acc_path.stem}.backup_*{acc_path.suffix}"))
    for old in backups[:-BACKUP_RETENTION]:
        old.unlink(missing_ok=True)
    return backup_path


def _ensure_plantilla_sheet(wb: Workbook, template_path: Path) -> Worksheet:
    if PLANTILLA_SHEET in wb.sheetnames:
        return wb[PLANTILLA_SHEET]
    src_wb = load_workbook(template_path)
    src_ws = src_wb.active
    dst = wb.create_sheet(PLANTILLA_SHEET)
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column):
        for src_cell in row:
            dst_cell = dst.cell(row=src_cell.row, column=src_cell.column)
            if not isinstance(src_cell, MergedCell):
                dst_cell.value = src_cell.value
            if src_cell.has_style:
                # Copia por componentes: los índices de _style no valen entre workbooks.
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.number_format = src_cell.number_format
    for rng in src_ws.merged_cells.ranges:
        dst.merge_cells(str(rng))
    for key, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst.column_dimensions[key].width = dim.width
    for key, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst.row_dimensions[key].height = dim.height
    dst.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dst.sheet_state = "hidden"
    src_wb.close()
    logger.info("Hoja %s creada dentro del acumulativo desde %s", PLANTILLA_SHEET, template_path.name)
    return dst


def _position_before_annual(wb: Workbook, sheet_name: str) -> None:
    if ANNUAL_SHEET not in wb.sheetnames:
        return
    current = wb.sheetnames.index(sheet_name)
    target = wb.sheetnames.index(ANNUAL_SHEET)
    if current > target:
        wb.move_sheet(sheet_name, offset=target - current)


def _pair_columns(pair: int) -> tuple[int, int]:
    name_col = FIRST_PAIR_COL + 2 * (pair - 1)
    return name_col, name_col + 1


def _detect_pairs(ws: Worksheet) -> int:
    pairs = 0
    for pair in range(1, 13):
        value = ws.cell(row=HEADER_TOP, column=_pair_columns(pair)[0]).value
        if isinstance(value, str) and "CARRERA" in value.upper():
            pairs = pair
        else:
            break
    if pairs == 0:
        raise CuadroGenerationError("La hoja plantilla del cuadro no tiene cabeceras de carrera (fila 17).")
    return pairs


def _spec_rows() -> list[int]:
    return [SPEC_FIRST_ROW + SPEC_BLOCK_ROWS * i for i in range(N_SPECIALIST_BLOCKS)]


def _adapt_race_pairs(ws: Worksheet, n_races: int) -> None:
    pairs = _detect_pairs(ws)
    # El merge del título (B13:O15 en plantilla) se reajusta al ancho real.
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row == TITLE_TOP and merged.min_col == 2:
            ws.unmerge_cells(str(merged))
            break
    if n_races < pairs:
        virgin_style = ws.cell(row=_VIRGIN_CELL[0], column=_VIRGIN_CELL[1])._style
        for pair in range(n_races + 1, pairs + 1):
            name_col, votes_col = _pair_columns(pair)
            for merged in list(ws.merged_cells.ranges):
                if (
                    merged.min_col >= name_col
                    and merged.max_col <= votes_col
                    and merged.min_row >= HEADER_TOP
                    and merged.max_row <= CONSENSUS_ROW
                ):
                    ws.unmerge_cells(str(merged))
            for row in range(TITLE_TOP, CONSENSUS_ROW + 1):
                if row == SUBTITLE_ROW:
                    continue
                for col in (name_col, votes_col):
                    cell = ws.cell(row=row, column=col)
                    cell.value = None
                    cell._style = copy(virgin_style)
    elif n_races > pairs:
        src_name_col, src_votes_col = _pair_columns(pairs)
        for pair in range(pairs + 1, n_races + 1):
            name_col, votes_col = _pair_columns(pair)
            for offset in range(2):
                src_letter = get_column_letter(src_name_col + offset)
                dst_letter = get_column_letter(name_col + offset)
                if src_letter in ws.column_dimensions and ws.column_dimensions[src_letter].width:
                    ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
            for row in range(TITLE_TOP, CONSENSUS_ROW + 1):
                if row == SUBTITLE_ROW:
                    continue
                for offset in range(2):
                    src_cell = ws.cell(row=row, column=src_name_col + offset)
                    dst_cell = ws.cell(row=row, column=name_col + offset)
                    dst_cell._style = copy(src_cell._style)  # mismo workbook: seguro
            ws.merge_cells(start_row=HEADER_TOP, start_column=name_col, end_row=HEADER_BOTTOM, end_column=votes_col)
            for r0 in _spec_rows():
                ws.merge_cells(start_row=r0, start_column=name_col, end_row=r0 + SPEC_BLOCK_ROWS - 1, end_column=name_col)
            ws.merge_cells(start_row=CONSENSUS_ROW, start_column=name_col, end_row=CONSENSUS_ROW, end_column=votes_col)
    last_votes_col = _pair_columns(n_races)[1]
    ws.merge_cells(start_row=TITLE_TOP, start_column=2, end_row=TITLE_BOTTOM, end_column=last_votes_col)


def _fill_sheet(ws: Worksheet, journey: Journey, settings: Settings) -> None:
    races_sorted = sorted(journey.races, key=lambda race: race.race_number)
    predictions_sorted = sorted(
        journey.predictions,
        key=lambda prediction: prediction.specialist.display_order if prediction.specialist else 99,
    )
    if len(predictions_sorted) != N_SPECIALIST_BLOCKS:
        raise CuadroGenerationError(
            f"Se esperaban {N_SPECIALIST_BLOCKS} pronósticos y hay {len(predictions_sorted)}."
        )

    # Limpiar datos de ejemplo de la plantilla (solo valores; anclas de merges).
    for pair in range(1, len(races_sorted) + 1):
        name_col, votes_col = _pair_columns(pair)
        for r0 in _spec_rows():
            ws.cell(row=r0, column=name_col).value = None
            for offset in range(SPEC_BLOCK_ROWS):
                ws.cell(row=r0 + offset, column=votes_col).value = None
        ws.cell(row=CONSENSUS_ROW, column=name_col).value = None
    for r0 in _spec_rows():
        ws.cell(row=r0, column=3).value = None

    ws.cell(row=SUBTITLE_ROW, column=1).value = f"{journey.venue} · {journey.date}"

    for index, race in enumerate(races_sorted, start=1):
        name_col, votes_col = _pair_columns(index)
        ws.cell(row=HEADER_TOP, column=name_col).value = f"{index}ª CARRERA"
        participants = {participant.number: participant.horse_name for participant in race.participants}
        for spec_index, prediction in enumerate(predictions_sorted):
            r0 = SPEC_FIRST_ROW + SPEC_BLOCK_ROWS * spec_index
            pick = next((p for p in prediction.picks if p.race_number == race.race_number), None)
            if pick is None:
                specialist = prediction.specialist.name if prediction.specialist else "?"
                raise CuadroGenerationError(f"{specialist} no tiene picks para la carrera {race.race_number}.")
            horse_name = participants.get(pick.pick_1)
            if horse_name is None:
                raise CuadroGenerationError(
                    f"El pick nº {pick.pick_1} de la carrera {race.race_number} no cruza con el partant oficial."
                )
            ws.cell(row=r0, column=name_col).value = f"{horse_name} ({pick.pick_1})"
            ws.cell(row=r0, column=votes_col).value = pick.pick_1
            ws.cell(row=r0 + 1, column=votes_col).value = pick.pick_2
            ws.cell(row=r0 + 2, column=votes_col).value = pick.pick_3

    for spec_index, prediction in enumerate(predictions_sorted):
        r0 = SPEC_FIRST_ROW + SPEC_BLOCK_ROWS * spec_index
        ws.cell(row=r0, column=3).value = prediction.specialist.name if prediction.specialist else ""

    pick_rows = [
        PickRow(race_number=pick.race_number, pick_1=pick.pick_1, pick_2=pick.pick_2, pick_3=pick.pick_3)
        for prediction in predictions_sorted
        for pick in prediction.picks
    ]
    consensus = calculate_consensus(
        pick_rows,
        pick_1_points=settings.consensus_pick_1_points,
        pick_2_points=settings.consensus_pick_2_points,
        pick_3_points=settings.consensus_pick_3_points,
    )
    for index, race in enumerate(races_sorted, start=1):
        top = consensus.get(race.race_number, [])
        ws.cell(row=CONSENSUS_ROW, column=_pair_columns(index)[0]).value = "-".join(str(n) for n in top[:3])


def _recolor(ws: Worksheet, theme: str, n_races: int) -> None:
    palette = CUADRO_PALETTES.get(theme)
    if palette is None:
        logger.warning("Tema sin paleta para el cuadro: %s (se mantiene el azul de plantilla)", theme)
        return
    dark = PatternFill("solid", fgColor=palette["dark"])
    light = PatternFill("solid", fgColor=palette["light"])
    last_col = _pair_columns(n_races)[1]
    for row in range(TITLE_TOP, CONSENSUS_ROW + 1):
        for col in range(1, last_col + 1):
            fill = ws.cell(row=row, column=col).fill
            if fill is None or fill.fill_type != "solid":
                continue
            color = fill.fgColor
            if color is None or color.type != "theme" or color.theme != 4:
                continue
            tint = color.tint or 0
            if tint < 0:
                ws.cell(row=row, column=col).fill = dark
            elif tint > 0.5:
                ws.cell(row=row, column=col).fill = light


def _attach_logo(ws: Worksheet, template_path: Path) -> None:
    try:
        with zipfile.ZipFile(template_path) as archive:
            data = archive.read(TEMPLATE_IMAGE_ENTRY)
        src_wb = load_workbook(template_path)
        src_images = getattr(src_wb.active, "_images", [])
        anchor = deepcopy(src_images[0].anchor) if src_images else "F5"
        src_wb.close()
        image = XLImage(BytesIO(data))
        image.anchor = anchor
        ws.add_image(image)
    except Exception:
        logger.warning("No se pudo insertar el logo en la hoja del cuadro; se continúa sin imagen.", exc_info=True)


def _verify(path: Path, expected_sheets: set[str], new_sheet: str) -> None:
    wb = load_workbook(path, read_only=True)
    try:
        names = set(wb.sheetnames)
        missing = expected_sheets - names
        if missing:
            raise CuadroGenerationError(f"Faltan hojas tras guardar: {sorted(missing)}")
        if new_sheet not in names:
            raise CuadroGenerationError(f"La hoja nueva {new_sheet} no está en el archivo guardado.")
        ws = wb[new_sheet]
        if ws["B17"].value != "ESPECIALISTAS":
            raise CuadroGenerationError("La hoja nueva no tiene la estructura esperada (B17).")
    finally:
        wb.close()


def _is_loadable(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=True)
        wb.close()
        return True
    except Exception:
        return False
