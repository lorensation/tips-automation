"""Genera el Excel de Tips rellenando la hoja de la sede en la plantilla oficial.

La plantilla nunca se recrea desde cero: se carga el workbook completo, se
conserva únicamente la hoja de la sede (con sus estilos, merges y anchos) y se
escriben los valores en las celdas existentes. Nunca se usan insert/delete de
filas o columnas porque openpyxl no desplaza merges ni estilos.
"""

import logging
from collections import Counter
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.enums import VENUE_TIPS_SHEETS
from app.models.journey import Journey

logger = logging.getLogger(__name__)

HEADER_ROW = 2
DATA_FIRST_ROW = 3
FOOTER_LABELS = (
    "FAVORITO DE LA JORNADA:",
    "ENTRENADOR DE LA JORNADA:",
    "JOCKEY DE LA JORNADA:",
)
# Celda alejada de la zona con datos: sirve como "estilo virgen" para limpiar.
_VIRGIN_CELL = (2, 40)


class TipsGenerationError(Exception):
    pass


def generate_tips_excel(journey: Journey, output_path: Path, template_path: Path) -> Path:
    if not template_path or not template_path.exists():
        raise TipsGenerationError(f"No se encuentra la plantilla de Tips: {template_path}")
    sheet_name = VENUE_TIPS_SHEETS.get(journey.venue)
    if sheet_name is None:
        raise TipsGenerationError(f"Sede sin hoja de Tips asociada: {journey.venue}")

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise TipsGenerationError(f"La plantilla no tiene la hoja '{sheet_name}'")
    ws = wb[sheet_name]
    for name in list(wb.sheetnames):
        if name != sheet_name:
            wb.remove(wb[name])
    wb.active = 0

    votes_by_race = _pick1_votes(journey)
    n_races = len(journey.races)
    sheet_groups = _detect_race_groups(ws)
    last_data_row = _detect_last_data_row(ws)
    footer_row = _detect_footer_row(ws, last_data_row)

    rows_available = last_data_row - DATA_FIRST_ROW + 1
    rows_needed = max((len(rows) for rows in votes_by_race.values()), default=0)
    if rows_needed > rows_available:
        delta = rows_needed - rows_available
        _shift_footer(ws, footer_row, delta)
        _extend_data_rows(ws, last_data_row, delta, max(n_races, sheet_groups))
        footer_row += delta
        last_data_row += delta

    _clear_example_data(ws, sheet_groups, last_data_row)

    if n_races > sheet_groups:
        for group in range(sheet_groups + 1, n_races + 1):
            _extend_race_group(ws, src_group=sheet_groups, dst_group=group, last_data_row=last_data_row)
    elif n_races < sheet_groups:
        for group in range(n_races + 1, sheet_groups + 1):
            _blank_race_group(ws, group, last_data_row)

    races_sorted = sorted(journey.races, key=lambda race: race.race_number)
    for index, race in enumerate(races_sorted, start=1):
        name_col, votes_col = _race_columns(index)
        ws.cell(row=HEADER_ROW, column=name_col).value = _ordinal_header(index)
        for offset, (horse_name, horse_number, count) in enumerate(votes_by_race.get(race.race_number, [])):
            row = DATA_FIRST_ROW + offset
            ws.cell(row=row, column=name_col).value = f"{horse_name} ({horse_number})"
            ws.cell(row=row, column=votes_col).value = count

    for offset, label in enumerate(FOOTER_LABELS):
        ws.cell(row=footer_row + offset, column=2).value = label

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(
        "Tips Excel generado para %s (%s carreras, hoja %s): %s",
        journey.venue, n_races, sheet_name, output_path,
    )
    return output_path


def _race_columns(group: int) -> tuple[int, int]:
    name_col = 2 + 3 * (group - 1)
    return name_col, name_col + 1


def _ordinal_header(group: int) -> str:
    return f"{group}ª CARRERA"


def _pick1_votes(journey: Journey) -> dict[int, list[tuple[str, int, int]]]:
    """Por carrera: [(nombre oficial, nº caballo, votos)] ordenado por (-votos, nº)."""
    result: dict[int, list[tuple[str, int, int]]] = {}
    for race in journey.races:
        counter: Counter[int] = Counter()
        for prediction in journey.predictions:
            for pick in prediction.picks:
                if pick.race_number == race.race_number:
                    counter[pick.pick_1] += 1
        participants = {participant.number: participant.horse_name for participant in race.participants}
        rows: list[tuple[str, int, int]] = []
        for horse_number, count in sorted(counter.items(), key=lambda item: (-item[1], item[0])):
            horse_name = participants.get(horse_number)
            if horse_name is None:
                raise TipsGenerationError(
                    f"El pick nº {horse_number} de la carrera {race.race_number} no cruza con el partant oficial."
                )
            rows.append((horse_name, horse_number, count))
        result[race.race_number] = rows
    return result


def _detect_race_groups(ws: Worksheet) -> int:
    groups = 0
    for group in range(1, 13):
        name_col, _ = _race_columns(group)
        value = ws.cell(row=HEADER_ROW, column=name_col).value
        if isinstance(value, str) and "CARRERA" in value.upper():
            groups = group
        else:
            break
    if groups == 0:
        raise TipsGenerationError("La hoja de la plantilla no tiene cabeceras de carrera en la fila 2.")
    return groups


def _detect_last_data_row(ws: Worksheet) -> int:
    """Última fila de la zona de datos: fila con borde inferior en la col B (cierre)."""
    for row in range(9, DATA_FIRST_ROW, -1):
        border = ws.cell(row=row, column=2).border
        if border.bottom and border.bottom.style:
            return row
    return 8


def _detect_footer_row(ws: Worksheet, last_data_row: int) -> int:
    for row in range(last_data_row + 1, last_data_row + 9):
        value = ws.cell(row=row, column=2).value
        if isinstance(value, str) and value.upper().startswith("FAVORITO"):
            return row
    return 10


def _clear_example_data(ws: Worksheet, sheet_groups: int, last_data_row: int) -> None:
    for group in range(1, sheet_groups + 1):
        name_col, votes_col = _race_columns(group)
        for row in range(DATA_FIRST_ROW, last_data_row + 1):
            ws.cell(row=row, column=name_col).value = None
            ws.cell(row=row, column=votes_col).value = None


def _blank_race_group(ws: Worksheet, group: int, last_data_row: int) -> None:
    name_col, votes_col = _race_columns(group)
    header_range = f"{get_column_letter(name_col)}{HEADER_ROW}:{get_column_letter(votes_col)}{HEADER_ROW}"
    for merged in list(ws.merged_cells.ranges):
        if str(merged) == header_range:
            ws.unmerge_cells(header_range)
            break
    virgin_style = ws.cell(row=_VIRGIN_CELL[0], column=_VIRGIN_CELL[1])._style
    for row in range(HEADER_ROW, last_data_row + 1):
        for col in (name_col, votes_col, votes_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell._style = copy(virgin_style)


def _extend_race_group(ws: Worksheet, src_group: int, dst_group: int, last_data_row: int) -> None:
    src_name_col, src_votes_col = _race_columns(src_group)
    dst_name_col, dst_votes_col = _race_columns(dst_group)
    for offset in range(3):
        src_letter = get_column_letter(src_name_col + offset)
        dst_letter = get_column_letter(dst_name_col + offset)
        if src_letter in ws.column_dimensions and ws.column_dimensions[src_letter].width:
            ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
    for row in range(HEADER_ROW, last_data_row + 1):
        for offset in range(3):
            src_cell = ws.cell(row=row, column=src_name_col + offset)
            dst_cell = ws.cell(row=row, column=dst_name_col + offset)
            dst_cell._style = copy(src_cell._style)
            dst_cell.value = None
    ws.merge_cells(
        start_row=HEADER_ROW, start_column=dst_name_col, end_row=HEADER_ROW, end_column=dst_votes_col
    )
    ws.cell(row=HEADER_ROW, column=dst_name_col).value = _ordinal_header(dst_group)


def _shift_footer(ws: Worksheet, footer_row: int, delta: int) -> None:
    virgin_style = ws.cell(row=_VIRGIN_CELL[0], column=_VIRGIN_CELL[1])._style
    footer_ranges = []
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= footer_row and merged.max_row <= footer_row + len(FOOTER_LABELS) - 1:
            footer_ranges.append((merged.min_row, merged.min_col, merged.max_row, merged.max_col))
            ws.unmerge_cells(str(merged))
    for row in range(footer_row, footer_row + len(FOOTER_LABELS)):
        for col in range(2, 6):
            src_cell = ws.cell(row=row, column=col)
            dst_cell = ws.cell(row=row + delta, column=col)
            dst_cell._style = copy(src_cell._style)
            dst_cell.value = src_cell.value
            src_cell.value = None
            src_cell._style = copy(virgin_style)
    for min_row, min_col, max_row, max_col in footer_ranges:
        ws.merge_cells(
            start_row=min_row + delta, start_column=min_col, end_row=max_row + delta, end_column=max_col
        )


def _extend_data_rows(ws: Worksheet, last_data_row: int, delta: int, total_groups: int) -> None:
    """Añade filas de datos manteniendo el alterno blanco/tinte y moviendo el cierre."""
    max_col = _race_columns(total_groups)[1] + 1
    parity_sources = {DATA_FIRST_ROW % 2: DATA_FIRST_ROW, (DATA_FIRST_ROW + 1) % 2: DATA_FIRST_ROW + 1}
    new_last_row = last_data_row + delta
    for row in range(last_data_row + 1, new_last_row + 1):
        src_row = parity_sources[row % 2]
        for col in range(2, max_col + 1):
            ws.cell(row=row, column=col)._style = copy(ws.cell(row=src_row, column=col)._style)
    # Mover el borde de cierre de la antigua última fila a la nueva.
    for col in range(2, max_col + 1):
        old_cell = ws.cell(row=last_data_row, column=col)
        parity_cell = ws.cell(row=parity_sources[last_data_row % 2], column=col)
        bottom_side = copy(old_cell.border.bottom)
        old_cell._style = copy(parity_cell._style)
        new_cell = ws.cell(row=new_last_row, column=col)
        border = new_cell.border
        new_cell.border = Border(
            left=copy(border.left), right=copy(border.right), top=copy(border.top), bottom=bottom_side
        )
    if last_data_row in ws.row_dimensions:
        for row in range(last_data_row + 1, new_last_row + 1):
            ws.row_dimensions[row].height = ws.row_dimensions[last_data_row].height
